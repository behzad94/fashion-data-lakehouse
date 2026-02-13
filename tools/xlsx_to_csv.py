from pathlib import Path
import csv
from openpyxl import load_workbook

def xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> None:

    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    sheet = wb.worksheets[0]

    csv_path.parent.mkdir(parents=True, exist_ok=True)

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(list(row))
    wb.close()


def main() -> None:
    xlsx_path = Path("data/source/online_retail/Online Retail.xlsx")
    csv_path = Path("data/source/online_retail/online_retail.csv")

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    
    xlsx_to_csv(xlsx_path, csv_path)
    print(f"Created: {csv_path}")


if __name__ == "__main__":
    main()